import pymupdf
from openpyxl import load_workbook
from datetime import datetime

file_path = 'default.xlsx'
workbook = load_workbook(filename=file_path)
doc = pymupdf.open("1.pdf")
try:
    sheet = workbook["List"]
    # Load a desired page. This works via 0-based numbers
    page = doc[0]  # this is the first page

    # Look for tables on this page and display the table count
    tabs = page.find_tables()
    print(f"{len(tabs.tables)} table(s) on {page}")
    tab = tabs[1]

    df = tab.to_pandas()  # convert to pandas DataFrame
    # df.to_excel(f"{doc.name}-{page.number}.xlsx")
    excel_row = 2
    for col_idx, (col_name, col_data) in enumerate(df.items()):
        if col_idx <= 1:
            continue
        for row_idx, cell in enumerate(col_data):
            if row_idx <= 3:
                continue
            print(f"Column {col_idx + 1} ({col_name}), Row {row_idx + 1}: {cell}")
            print(df.iloc[row_idx, col_idx])
            if (row_idx -4) % 2 == 0:
                if df.iloc[row_idx, col_idx] not in ["None", "", "SKIP"] and df.iloc[row_idx + 1, col_idx] not in ["None", "", "SKIP"]:
                    print(df.iloc[0, col_idx])
                    sheet.cell(row=excel_row, column=1, value=str(excel_row - 1))
                    sheet.cell(row=excel_row, column=2, value="KRKUV")
                    sheet.cell(row=excel_row, column=3, value="K-line Ro-Ro")
                    sheet.cell(row=excel_row, column=4, value=df.iloc[0, col_idx].replace("\n", " "))
                    sheet.cell(row=excel_row, column=5, value=df.iloc[1, col_idx])
                    sheet.cell(row=excel_row, column=6, value=df.iloc[row_idx, 0])
                    date_object = datetime.strptime(df.iloc[row_idx, col_idx], "%d-%b")
                    date_object = date_object.replace(year=2024)
                    formatted_date = date_object.strftime("%Y-%m-%d")
                    sheet.cell(row=excel_row, column=7, value=formatted_date)
                    date_object = datetime.strptime(df.iloc[row_idx + 1, col_idx], "%d-%b")
                    date_object = date_object.replace(year=2024)
                    formatted_date = date_object.strftime("%Y-%m-%d")
                    sheet.cell(row=excel_row, column=10, value=formatted_date)
                    excel_row += 1
    # We will see a message like "1 table(s) on page 0 of input.pdf"
finally:
    workbook.save("pdfoutput.xlsx")
