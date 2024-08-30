from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import sys


def main():
    driver = webdriver.Chrome()
    file_path = 'default.xlsx'  # Path to your existing Excel file
    workbook = load_workbook(filename=file_path)
    default_arg1 = "germany"
    default_arg2 = "china"
    # Check if enough arguments are provided
    if len(sys.argv) > 1:
        arg1 = sys.argv[1]
    else:
        arg1 = default_arg1

    if len(sys.argv) > 2:
        arg2 = sys.argv[2]
    else:
        arg2 = default_arg2
    try:
        sheet = workbook["List"]
        print("Here the sample test case will be started")
        driver.maximize_window()
        driver.get("https://m.eclipsocean.com/ek/otsd/homepage/01_ShippingService/otsdPortSchedule.do")
        driver.find_element(By.XPATH, '//*[@id="wrap"]/div[1]/div/div[1]/span[2]/span[1]/span/span[2]').click()
        driver.find_element(By.XPATH, '/html/body/span/span/span[1]/input').send_keys(arg1 + Keys.ENTER)
        driver.find_element(By.XPATH, '//*[@id="wrap"]/div[1]/div/div[1]/span[4]/span[1]/span/span[2]').click()
        driver.find_element(By.XPATH, '/html/body/span/span/span[1]/input').send_keys(arg2 + Keys.ENTER)
        driver.find_element(By.XPATH, '//*[@id="s_btn"]').click()
        # waiting until data fill
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#data_area > *"))
        )

        # Locate the <div> with id 'data_area'
        data_area_element = driver.find_element(By.ID, 'data_area')

        # Extract and print the text content of the <tbody> within the data_area
        tbody_element = data_area_element.find_element(By.TAG_NAME, 'tbody')

        # Optionally, you can loop through rows and columns if needed
        rows = tbody_element.find_elements(By.TAG_NAME, 'tr')
        excel_row = 2
        vessel_txts=["", ""]
        for row in rows:
            cols = row.find_elements(By.TAG_NAME, 'td')
            if len(cols) == 0:
                continue
            if len(cols[0].text.split(" V-")) == 2:
                vessel_txts = cols[0].text.split(" V-")
                sheet.cell(row=excel_row, column=1, value=str(excel_row - 1))
                sheet.cell(row=excel_row, column=2, value="KRKUV")
                sheet.cell(row=excel_row, column=3, value="Eukor")
                sheet.cell(row=excel_row, column=4, value=vessel_txts[0] if len(vessel_txts) == 2 else "")
                sheet.cell(row=excel_row, column=5, value=("V-" + vessel_txts[1]) if len(vessel_txts) == 2 else "")
                sheet.cell(row=excel_row, column=6, value=cols[1].text)
                sheet.cell(row=excel_row, column=7, value=cols[2].text)
                sheet.cell(row=excel_row, column=10, value=cols[2].text)
                excel_row += 1
                sheet.cell(row=excel_row, column=1, value=str(excel_row - 1))
                sheet.cell(row=excel_row, column=2, value="KRKUV")
                sheet.cell(row=excel_row, column=3, value="Eukor")
                sheet.cell(row=excel_row, column=4, value=vessel_txts[0] if len(vessel_txts) == 2 else "")
                sheet.cell(row=excel_row, column=5, value=("V-" + vessel_txts[1]) if len(vessel_txts) == 2 else "")
                sheet.cell(row=excel_row, column=6, value=cols[5].text)
                sheet.cell(row=excel_row, column=7, value=cols[6].text)
                sheet.cell(row=excel_row, column=10, value=cols[6].text)
                excel_row += 1
            else:
                sheet.cell(row=excel_row, column=1, value=str(excel_row - 1))
                sheet.cell(row=excel_row, column=2, value="KRKUV")
                sheet.cell(row=excel_row, column=3, value="Eukor")
                sheet.cell(row=excel_row, column=4, value=vessel_txts[0] if len(vessel_txts) == 2 else "")
                sheet.cell(row=excel_row, column=5, value=("V-" + vessel_txts[1]) if len(vessel_txts) == 2 else "")
                sheet.cell(row=excel_row, column=6, value=cols[5].text)
                sheet.cell(row=excel_row, column=7, value=cols[6].text)
                sheet.cell(row=excel_row, column=10, value=cols[6].text)
                excel_row += 1
    finally:
        workbook.save("output.xlsx")
        driver.close()
    print("sample test case successfully completed")


if __name__ == "__main__":
    main()
